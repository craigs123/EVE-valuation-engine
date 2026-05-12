"""Generate mangrove_test_calculator.xlsx — a spreadsheet that mirrors
the live mangrove ESVD coefficients with input cells for Area, Intactness
and Regional adjustment so you can manually cross-check EVE's totals.

Run from repo root: python scripts/generate_mangrove_calculator.py
"""
import os
import sys

# Make sibling utils/ importable when running this script directly
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _REPO_ROOT)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.precomputed_esvd_coefficients import PrecomputedESVDCoefficients
from utils.project_indicators_seed import DEFAULT_INDICATORS
from utils.teeb_slug_map import TEEB_TO_CALC_KEY, WEIGHT_LOOKUP, HD_RELATIONSHIP


calc = PrecomputedESVDCoefficients()
mangrove_coeffs = calc.get_ecosystem_coefficients('Mangroves')


# Order matches the service_categories grouping in the calc engine
service_order = [
    ('Provisioning', [
        ('food',                  'Food'),
        ('water',                 'Water'),
        ('raw_materials',         'Raw materials'),
        ('genetic_resources',     'Genetic resources'),
        ('medicinal_resources',   'Medicinal resources'),
        ('ornamental_resources',  'Ornamental resources'),
    ]),
    ('Regulating', [
        ('pollution',             'Air quality regulation'),
        ('climate',               'Climate regulation'),
        ('extreme_events',        'Moderation of extreme events'),
        ('water_regulation',      'Water flow regulation'),
        ('water_purification',    'Waste treatment / water purification'),
        ('erosion',               'Erosion prevention'),
        ('soil_formation',        'Soil formation / nutrient cycling'),
        ('pollination',           'Pollination'),
        ('biological_control',    'Biological control'),
    ]),
    ('Supporting', [
        ('nursery_services',      'Maintenance of life cycles (nursery)'),
        ('habitat',               'Habitat / Genetic diversity'),
    ]),
    ('Cultural', [
        ('aesthetic_value',       'Aesthetic information'),
        ('recreation',            'Recreation & tourism'),
        ('cultural',              'Inspiration / culture'),
        ('spiritual_value',       'Spiritual experience'),
        ('primary_production',    'Cognitive development'),
    ]),
]


wb = Workbook()
ws = wb.active
ws.title = 'Mangrove Calculator'

BOLD_WHITE = Font(bold=True, color='FFFFFF')
GREEN = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
LIGHT = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
THIN = Side(border_style='thin', color='B0BEC5')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal='center', vertical='center')

# Header
ws['A1'] = 'EVE Mangrove Ecosystem Services — Test Calculator'
ws['A1'].font = Font(bold=True, size=14, color='1B5E20')
ws.merge_cells('A1:G1')

ws['A3'] = 'Inputs (yellow cells — edit me)'
ws['A3'].font = Font(bold=True, color='594400')
ws.merge_cells('A3:G3')

ws['A4'] = 'Area (ha)'
ws['B4'] = 1000
ws['B4'].fill = INPUT_FILL
ws['B4'].border = BORDER
ws['B4'].number_format = '#,##0.00'

ws['A5'] = 'Intactness multiplier (0.0–1.0)'
ws['B5'] = 1.00
ws['B5'].fill = INPUT_FILL
ws['B5'].border = BORDER
ws['B5'].number_format = '0.00'

ws['A6'] = 'Regional adjustment factor (0.4–2.5)'
ws['B6'] = 1.00
ws['B6'].fill = INPUT_FILL
ws['B6'].border = BORDER
ws['B6'].number_format = '0.00'

ws['A7'] = 'Sundarbans (India) regional factor'
ws['B7'] = 0.51
ws['B7'].font = Font(italic=True, color='6B7280')
ws['B7'].number_format = '0.00'
ws['C7'] = '← copy into B6 to test the Sundarbans Mangrove test area'
ws['C7'].font = Font(italic=True, color='6B7280')

# Table header
header_row = 9
headers = ['Service', 'Category', 'Coefficient (Int$/ha/yr)',
           '× Area', '× Intactness', '× Regional',
           'Contribution (Int$/yr)']
for col_idx, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col_idx, value=h)
    c.font = BOLD_WHITE
    c.fill = GREEN
    c.alignment = CENTER
    c.border = BORDER

# Service rows
row = header_row + 1
first_row = row
for category, services in service_order:
    for key, display_name in services:
        coeff = mangrove_coeffs.get(key, 0.0)
        ws.cell(row=row, column=1, value=display_name).border = BORDER
        ws.cell(row=row, column=2, value=category).border = BORDER
        ws.cell(row=row, column=3, value=coeff).border = BORDER
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        # Display the live multiplier values (echo from input cells)
        ws.cell(row=row, column=4, value='=$B$4').border = BORDER
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value='=$B$5').border = BORDER
        ws.cell(row=row, column=5).number_format = '0.00'
        ws.cell(row=row, column=6, value='=$B$6').border = BORDER
        ws.cell(row=row, column=6).number_format = '0.00'
        # Contribution
        contribution = ws.cell(row=row, column=7,
                               value=f'=C{row}*$B$4*$B$5*$B$6')
        contribution.border = BORDER
        contribution.number_format = '"$"#,##0.00'
        row += 1
last_row = row - 1

# TOTAL row
ws.cell(row=row, column=1, value='TOTAL ANNUAL VALUE').font = Font(bold=True, size=12)
ws.cell(row=row, column=1).fill = LIGHT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
total = ws.cell(row=row, column=7, value=f'=SUM(G{first_row}:G{last_row})')
total.font = Font(bold=True, size=12, color='1B5E20')
total.fill = LIGHT
total.border = BORDER
total.number_format = '"$"#,##0.00'

# Per-hectare
row += 1
ws.cell(row=row, column=1,
        value='Total per hectare (Int$/ha/yr)').font = Font(italic=True, color='6B7280')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
per_ha = ws.cell(row=row, column=7, value=f'=G{row - 1}/$B$4')
per_ha.number_format = '"$"#,##0.00'
per_ha.font = Font(italic=True, color='6B7280')

# Notes
row += 2
notes = [
    'Notes:',
    '• Coefficients sourced from TEEB analysis (ESVD 2.4.2 medians with '
    'literature substitutions where ESVD data is sparse).',
    '• Multipliers in this sheet are applied uniformly across all '
    'services — i.e. it mirrors the legacy BBI calculation mode.',
    '• To test the Sundarbans Mangrove test area in EVE: '
    'set Area=1000, Intactness=1.0, Regional=0.51 → expected $15,764,610/yr.',
    '• Indicator-driven (non-uniform) multipliers are NOT modelled in this '
    'sheet — for those, use the EVE results page sub-service breakdown.',
    '• Coefficient source: utils/precomputed_esvd_coefficients.py · '
    "ecosystem-dict key 'mangroves' · base sum = $30,911/ha/yr.",
]
for note in notes:
    c = ws.cell(row=row, column=1, value=note)
    if note.startswith('Notes:'):
        c.font = Font(bold=True)
    else:
        c.font = Font(color='6B7280', size=10)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

# Column widths
widths = {1: 38, 2: 14, 3: 24, 4: 11, 5: 13, 6: 12, 7: 22}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Sheet 2: Indicator → sub-service mapping ────────────────────────────────
ws2 = wb.create_sheet(title='Indicator Mapping')

PRIMARY_FILL = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # green
SECONDARY_FILL = PatternFill(start_color='AED581', end_color='AED581', fill_type='solid')  # lighter green
HD_FILL = PatternFill(start_color='FB8C00', end_color='FB8C00', fill_type='solid')  # amber
SUB_HEADER_FILL = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')

# Title + notes
ws2['A1'] = 'Indicator → Sub-service mapping (Mangrove)'
ws2['A1'].font = Font(bold=True, size=14, color='1B5E20')
ws2.merge_cells('A1:Z1')

ws2['A2'] = (
    'Each cell shows how strongly an indicator drives the multiplier for that '
    'sub-service. Green = primary (weight 1.0), light-green = secondary '
    '(weight 0.5), amber = cross-cutting (HD applies to all).'
)
ws2['A2'].font = Font(color='6B7280', size=10, italic=True)
ws2['A2'].alignment = Alignment(wrap_text=True)
ws2.merge_cells('A2:Z2')
ws2.row_dimensions[2].height = 36

# Build the flat list of mangrove sub-service keys (same order as Calculator tab)
sub_service_keys = []
sub_service_labels = []
for category, services in service_order:
    for key, display_name in services:
        sub_service_keys.append(key)
        sub_service_labels.append(display_name)

# Build a lookup of (indicator slug, calc key) → relationship type
# When an indicator double-maps to one calc key (e.g. habitat_for_species P +
# genetic_diversity S both → 'habitat'), take the STRONGEST relationship,
# matching what _compute_pure() does in the calc engine.
def _rank(rel):
    return WEIGHT_LOOKUP.get(rel, 0.0) if rel != HD_RELATIONSHIP else 0.0

indicator_to_sub_service: dict[tuple[str, str], str] = {}
for ind in DEFAULT_INDICATORS:
    slug = ind['slug']
    sw = ind.get('service_weights') or {}
    for teeb_slug, rel in sw.items():
        calc_key = TEEB_TO_CALC_KEY.get(teeb_slug)
        if calc_key is None:
            continue
        # Track best relationship for (indicator, calc_key)
        existing = indicator_to_sub_service.get((slug, calc_key))
        if existing is None or (rel == HD_RELATIONSHIP) or (_rank(rel) > _rank(existing)):
            indicator_to_sub_service[(slug, calc_key)] = rel

# Header row
header_row_2 = 4
ws2.cell(row=header_row_2, column=1, value='Indicator').font = BOLD_WHITE
ws2.cell(row=header_row_2, column=1).fill = GREEN
ws2.cell(row=header_row_2, column=1).alignment = CENTER
ws2.cell(row=header_row_2, column=1).border = BORDER
for col_idx, label in enumerate(sub_service_labels, start=2):
    c = ws2.cell(row=header_row_2, column=col_idx, value=label)
    c.font = BOLD_WHITE
    c.fill = GREEN
    c.alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=60, wrap_text=True)
    c.border = BORDER
ws2.row_dimensions[header_row_2].height = 140

# Indicator rows
row2 = header_row_2 + 1
for ind in DEFAULT_INDICATORS:
    slug = ind['slug']
    code = ind.get('code') or '?'
    name = ind.get('name') or slug
    is_mandatory = bool(ind.get('is_mandatory'))
    label_cell = ws2.cell(row=row2, column=1, value=f"{code}: {name}" + (' (Required)' if is_mandatory else ''))
    label_cell.font = Font(bold=True, color='594400' if is_mandatory else '1B5E20')
    label_cell.fill = SUB_HEADER_FILL
    label_cell.border = BORDER
    for col_idx, key in enumerate(sub_service_keys, start=2):
        rel = indicator_to_sub_service.get((slug, key))
        cell = ws2.cell(row=row2, column=col_idx)
        cell.border = BORDER
        cell.alignment = CENTER
        if rel == 'primary':
            cell.value = 'P'
            cell.fill = PRIMARY_FILL
            cell.font = Font(bold=True, color='FFFFFF')
        elif rel == 'secondary':
            cell.value = 'S'
            cell.fill = SECONDARY_FILL
            cell.font = Font(bold=True, color='1B5E20')
        elif rel == HD_RELATIONSHIP:
            cell.value = 'C'
            cell.fill = HD_FILL
            cell.font = Font(bold=True, color='FFFFFF')
        # else: leave blank (no relationship)
    row2 += 1

# Legend
row2 += 1
ws2.cell(row=row2, column=1, value='Legend').font = Font(bold=True)
row2 += 1
ws2.cell(row=row2, column=1, value='P = Primary (weight 1.0)')
ws2.cell(row=row2, column=2).fill = PRIMARY_FILL
ws2.cell(row=row2, column=2, value='P').font = Font(bold=True, color='FFFFFF')
ws2.cell(row=row2, column=2).alignment = CENTER
row2 += 1
ws2.cell(row=row2, column=1, value='S = Secondary (weight 0.5)')
ws2.cell(row=row2, column=2).fill = SECONDARY_FILL
ws2.cell(row=row2, column=2, value='S').font = Font(bold=True, color='1B5E20')
ws2.cell(row=row2, column=2).alignment = CENTER
row2 += 1
ws2.cell(row=row2, column=1, value='C = Cross-cutting (HD multiplier applied to all)')
ws2.cell(row=row2, column=2).fill = HD_FILL
ws2.cell(row=row2, column=2, value='C').font = Font(bold=True, color='FFFFFF')
ws2.cell(row=row2, column=2).alignment = CENTER
row2 += 1
ws2.cell(row=row2, column=1,
         value='Blank = indicator does not affect this sub-service').font = Font(italic=True, color='6B7280')

# Column widths
ws2.column_dimensions['A'].width = 36
for col_idx in range(2, 2 + len(sub_service_labels)):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 6

out_path = os.path.join(_REPO_ROOT, 'Mangrove_Test.xlsx')
wb.save(out_path)

# Sanity-print the totals so we know the sheet's formulas should match
base_sum = sum(mangrove_coeffs.values())
print(f'Wrote {out_path}')
print(f'Base sum per ha:                ${base_sum:,.2f}/ha/yr')
print(f'A=1000, I=1.0, R=1.0 :          ${base_sum * 1000 * 1.0 * 1.00:,.2f}')
print(f'A=1000, I=1.0, R=0.51 (Sund.):  ${base_sum * 1000 * 1.0 * 0.51:,.2f}')
print(f'A=1000, I=0.7, R=2.5 (max reg): ${base_sum * 1000 * 0.7 * 2.50:,.2f}')
